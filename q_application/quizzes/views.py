from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Avg
from .models import CustomUser, Subject, Question, Result, UserAnswer
from .forms import AddStudentForm, RemoveStudentForm, SubjectForm
import datetime
from django.http import HttpResponse
from openpyxl import load_workbook
from django.contrib import messages

def front_page(request):
    return render(request, 'front_page.html')

# User Login View
def login_view(request):
    if request.method == 'POST':
        # Get the user input
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Try to authenticate using either username or student_id - Username for admin and student_id for students
        user = None
        if CustomUser.objects.filter(username=username).exists():
            user = authenticate(request, username=username, password=password)
        elif CustomUser.objects.filter(student_id=username).exists():
            user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            if user.role == 'admin':
                return redirect('admin_dashboard')
            else:
                return redirect('student_dashboard')
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'login.html')

@login_required
def admin_dashboard(request):
    total_students = CustomUser.objects.filter(role='student').count()
    active_subjects = Subject.objects.count()
    total_tests = Result.objects.count()
    avg_correct_answers = Result.objects.aggregate(avg_correct=Avg('correct_answers'))['avg_correct'] or 0

    context = {
        'total_students': total_students,
        'active_subjects': active_subjects,
        'total_tests': total_tests,
        'avg_correct_answers': round(avg_correct_answers, 2)
    }
    return render(request, 'admin_dashboard.html', context)

@login_required
def manage_users(request):
    add_form = AddStudentForm()
    remove_form = RemoveStudentForm()
    students = CustomUser.objects.filter(role='student')

    if request.method == "POST":
        if "add_student" in request.POST:
            add_form = AddStudentForm(request.POST)
            if add_form.is_valid():
                student = add_form.save(commit=False)
                student.role = 'student'
                student.set_password(add_form.cleaned_data['password'])  # Hash password
                student.save()
                return redirect('manage_users')

        elif "remove_student" in request.POST:
            remove_form = RemoveStudentForm(request.POST)
            if remove_form.is_valid():
                student_id = remove_form.cleaned_data['student_id']
                CustomUser.objects.filter(student_id=student_id, role='student').delete()
                return redirect('manage_users')

    return render(request, 'admin_manage_users.html', {
        'students': students,
        'add_form': add_form,
        'remove_form': remove_form
    })


@login_required
def add_subjects(request):
    if request.method == 'POST':
        form = SubjectForm(request.POST)
        mcq_easy = request.FILES.get('mcq_easy')
        mcq_medium = request.FILES.get('mcq_medium')
        mcq_hard = request.FILES.get('mcq_hard')
        oneword_easy = request.FILES.get('oneword_easy')
        oneword_medium = request.FILES.get('oneword_medium')
        oneword_hard = request.FILES.get('oneword_hard')

        if form.is_valid() and mcq_easy and mcq_medium and mcq_hard and oneword_easy and oneword_medium and oneword_hard:
            subject = form.save()
            try:
                # Handle file uploads
                process_excel_file(mcq_easy, subject, 'mcq', 'easy')
                process_excel_file(mcq_medium, subject, 'mcq', 'medium')
                process_excel_file(mcq_hard, subject, 'mcq', 'hard')
                process_excel_file(oneword_easy, subject, 'oneword', 'easy')
                process_excel_file(oneword_medium, subject, 'oneword', 'medium')
                process_excel_file(oneword_hard, subject, 'oneword', 'hard')

                messages.success(request, "Subject and questions added successfully!")
                return redirect('manage_subjects')
            except Exception as e:
                messages.error(request, f"Error processing files: {e}")
        else:
            messages.error(request, "Please provide all required files and valid inputs.")
    else:
        form = SubjectForm()

    return render(request, 'admin_add_subjects.html', {'form': form})
def process_excel_file(file, subject, question_type, difficulty):
    workbook = load_workbook(file)
    sheet = workbook.active

    if question_type == 'mcq':
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            Question.objects.create(
                subject=subject,
                question_text=row[1],
                question_type='mcq',
                difficulty=difficulty,
                options={
                    'option1': row[2],
                    'option2': row[3],
                    'option3': row[4],
                    'option4': row[5],
                },
                correct_answer=row[6]
            )
    elif question_type == 'oneword':
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            Question.objects.create(
                subject=subject,
                question_text=row[1],
                question_type='oneword',
                difficulty=difficulty,
                correct_answer=row[2]
            )

def view_results(request):
    query = request.GET.get('q', '')  # Capture the search query
    results = Result.objects.all().select_related('user', 'subject')

    # Apply filtering if a search query is provided
    if query:
        results = results.filter(user__username__icontains=query) | results.filter(user__student_id__icontains=query)

    context = {
        'results': results,
        'search_query': query,
    }
    return render(request, 'admin_view_results.html', context)



def search_student(request):
    if request.method == 'GET':
        search_query = request.GET.get('q', '').strip()  # Get search input
        students = []
        if search_query:
            # Filter students by username or student_id where role is 'student'
            students = CustomUser.objects.filter(
                role='student',
                student_id__icontains=search_query
            ) | CustomUser.objects.filter(
                role='student',
                username__icontains=search_query
            )

            # If no students are found, show a message
            if not students.exists():
                messages.info(request, "User not found.")
        
        else:
            messages.warning(request, "Please enter a search term.")
        
        return render(request, 'admin_view_results.html', {'students': students, 'search_query': search_query})

def quiz_selection(request):
    if request.method == 'POST':
        subject_id = request.POST['subject']
        quiz_type = request.POST['quiz_type']
        
        if quiz_type == "":
            message = "Please select a question type!"
            return render(request, 'quiz_selection.html', {'message': message, 'subjects': Subject.objects.all()})
        
        return redirect('quiz_page', subject_id=subject_id)

    return render(request, 'quiz_selection.html', {'subjects': Subject.objects.all()})


def quiz_page(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    user = request.user
    question_index = request.session.get('question_index', 0)
    quiz_type = request.GET.get('quiz_type')

    # Initialize session data if not exists
    if 'user_answers' not in request.session:
        request.session['user_answers'] = []
    
    # Set initial difficulty
    difficulty = 'medium'
    if question_index > 0:
        user_answers = request.session.get('user_answers', [])
        if user_answers:
            last_answer = user_answers[-1]
            difficulty = 'easy' if last_answer.get('correct') else 'hard'

    if request.method == 'POST':
        # Get the current question's ID from the hidden field
        current_question_id = request.POST.get('question_id')
        current_question = Question.objects.get(id=current_question_id)
        
        answer_selected = request.POST.get('answer')
        correct = str(answer_selected).strip().lower() == str(current_question.correct_answer).strip().lower()
        
        start_time = float(request.POST.get('start_time'))
        end_time = datetime.datetime.now().timestamp()
        time_taken = end_time - start_time

        # Store answer data
        answer_data = {
            'question_id': current_question.id,
            'answer_selected': answer_selected,
            'correct': correct,
            'difficulty': difficulty,
            'time_taken': time_taken
        }
        user_answers = request.session.get('user_answers', [])
        user_answers.append(answer_data)
        request.session['user_answers'] = user_answers
        request.session.modified = True

        question_index += 1
        request.session['question_index'] = question_index

        # Handle quiz completion
        if question_index >= 10:
            # Calculate results
            correct_answers = sum(1 for answer in user_answers if answer['correct'])
            times = [answer['time_taken'] for answer in user_answers]

            # Save individual answers
            for answer_data in user_answers:
                UserAnswer.objects.create(
                    user=user,
                    question=Question.objects.get(id=answer_data['question_id']),
                    answer_selected=answer_data['answer_selected'],
                    correct=answer_data['correct'],
                    difficulty=answer_data['difficulty'],
                    time_taken=answer_data['time_taken']
                )

            # Calculate marks and grade
            marks = correct_answers
            grade = 'A' if marks >= 80 else 'B' if marks >= 70 else 'C' if marks >= 50 else 'F'

            # Save result
            Result.objects.create(
                user=user,
                subject=subject,
                quiz_type=quiz_type,
                correct_answers=correct_answers,
                total_questions=10,
                marks=marks,
                grade=grade,
                times=times
            )

            # Clear session
            request.session['user_answers'] = []
            request.session['question_index'] = 0
            if 'start_time' in request.session:
                del request.session['start_time']
            request.session.modified = True

            return redirect('student_dashboard')

    # Get next question
    questions = Question.objects.filter(
        subject=subject,
        question_type=quiz_type,
        difficulty=difficulty
    ).order_by('?')

    if not questions.exists():
        return HttpResponse("Not enough questions for the selected difficulty level.", status=400)

    question = questions[0]
    
    # Set start time for first question
    if 'start_time' not in request.session:
        request.session['start_time'] = datetime.datetime.now().timestamp()

    return render(request, 'quiz_page.html', {
        'question': question,
        'subject': subject,
        'user': user,
        'start_time': request.session.get('start_time'),
        'quiz_type': quiz_type,
        'question_index': question_index,
    })

@login_required
def manage_subjects(request):
    subjects = Subject.objects.all()
    if request.method == 'POST':
        subject_id = request.POST.get('remove_subject')  # Fix to match the form input name
        Subject.objects.filter(id=subject_id).delete()
    return render(request, 'admin_manage_subjects.html', {'subjects': subjects})

@login_required
def student_dashboard(request):
    student_id = request.user.student_id  # Assuming the user is the student
    return render(request, 'student_dashboard.html', {'student_id': student_id})

@login_required
def student_settings(request):
    # Ensure only students can access this view
    if request.user.role != 'student':
        messages.error(request, "You are not authorized to access this page.")
        return redirect('student_dashboard')  # Redirect non-students to their dashboard
    student_id = request.user.student_id
    if request.method == 'POST':
        # Update student profile details
        name = request.POST.get('name')
        age = request.POST.get('age')
        school = request.POST.get('school')
        class_name = request.POST.get('class_name')

        # Update the user's details
        user = request.user
        user.name = name
        user.age = age
        user.school = school
        user.class_name = class_name
        user.save()

        messages.success(request, "Your profile has been updated successfully!")
        return redirect('student_settings')

    return render(request, 'student_settings.html', {
        'user': request.user,
        'student_id' : student_id,
    })

@login_required
def user_results(request):
    # Fetch all results for the logged-in student, ordered by most recent
    results = Result.objects.filter(user=request.user).order_by('-id')
    
    # Process each result to add additional computed fields
    for result in results:
        result.incorrect_answers = result.total_questions - result.correct_answers
        # Marks are already out of 100, no need to format
        result.formatted_marks = int(result.marks)  # Convert to integer for clean display
        # Calculate success rate
        result.success_rate = (result.correct_answers / result.total_questions) * 100
        # Calculate average time per question
        if result.times:
            avg_time = sum(result.times) / len(result.times)
            result.average_time = "{:.2f}".format(avg_time)
        else:
            result.average_time = "N/A"

    context = {
        'results': results,
        'student_id': request.user.student_id,
        'has_results': results.exists(),
    }
    
    return render(request, 'user_results.html', context)


def logout_view(request):
    logout(request)  # Logs out the user
    return redirect('login')  # Redirects to the login page after logout